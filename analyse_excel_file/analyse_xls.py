from typing import Dict, Any

import openpyxl

path = './'

xl_file = openpyxl.load_workbook(path)
xl_file_list = xl_file['Sheet1']
products_per_supplier: dict[Any, Any] = {}
product_inventory_less_than_ten = {}
total_inventory_value_per_company = {}


def list_company_product_count():
    for each_row in range(2, xl_file_list.max_row + 1):
        supplier_name = xl_file_list.cell(each_row, 4).value
        if supplier_name in products_per_supplier:
            products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1

        else:
            print("adding a new supplier")
            products_per_supplier[supplier_name] = 1


def list_product_inventory_less_than_ten():
    for each_row in range(2, xl_file_list.max_row + 1):
        if xl_file_list.cell(each_row, 2).value < 10:
            less_than_ten_value = int(xl_file_list.cell(each_row, 2).value)
            less_than_ten_product = int(xl_file_list.cell(each_row, 1).value)
            product_inventory_less_than_ten[less_than_ten_product] = less_than_ten_value
        else:
            pass


def list_company_inventory_value():
    for each_row in range(2, xl_file_list.max_row + 1):
        supplier_name = xl_file_list.cell(each_row, 4).value
        if supplier_name in total_inventory_value_per_company:
            product_value = xl_file_list.cell(each_row, 3).value * xl_file_list.cell(each_row, 2).value
            new_total_value = total_inventory_value_per_company.get(supplier_name) + product_value
            total_inventory_value_per_company[supplier_name] = float(new_total_value)
        else:
            print("adding a new supplier")
            product_value = xl_file_list.cell(each_row, 3).value * xl_file_list.cell(each_row, 2).value
            total_inventory_value_per_company[supplier_name] = product_value


def write_inventory_value_per_product_to_file():
    for each_row in range(2, xl_file_list.max_row + 1):
        supplier_name = xl_file_list.cell(each_row, 4).value
        product_value = xl_file_list.cell(each_row, 3).value * xl_file_list.cell(each_row, 2).value
        xl_file_list.cell(each_row, 5).value = product_value


list_company_product_count()
list_product_inventory_less_than_ten()
list_company_inventory_value()
write_inventory_value_per_product_to_file()
print(f"These are the number of products per supplier \n {products_per_supplier}")
print(f"These are the products that are less than 10 \n {product_inventory_less_than_ten}")
print(f"Below is company inventory value in total \n {total_inventory_value_per_company}")
xl_file.save(path)
